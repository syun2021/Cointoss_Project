
import PySimpleGUI as sg
import numpy as np
import openpyxl
import glob
import pandas as pd
import tkinter
import sys
from tkinter import messagebox
import random
import sys
import matplotlib.pyplot as plt

#初期値
day=[
    "月曜日",
    "火曜日",
    "水曜日",
    "木曜日",
    "金曜日"
]
text=[
    "表",
    "裏",
    "勝ち",
    "負け"
]
global text2
text2=[
    "lot数",
    0.5,
    "指値",
    60,
    "逆指値",
    -50
]
#global text2
chk={}
bln={}

a=False

root=tkinter.Tk()
root.geometry('400x400')
root.title("コイントス結果")

for i in range(len(day)):
    tkinter.Label(root,text=day[i]).place(x=20,y=(i+1)*50)
for i in range(len(text)):
    tkinter.Label(root,text=text[i]).place(x=(i+2)*50,y=20)

for i in range(len(text)):
    for j in range(len(day)):
        bln[i*len(day)+j]=tkinter.BooleanVar()
        chk[i*len(day)+j]=tkinter.Checkbutton(root,variable=bln[i*len(day)+j],text="")
        chk[i*len(day)+j].place(x=(i+2)*50,y=(j+1)*50)

weeks=tkinter.IntVar()
edit=tkinter.ttk.Entry(root,textvariable=weeks,width=2)
edit.place(x=20,y=20)
tkinter.Label(root,text="週目").place(x=55,y=20)
tkinter.Button(root,text=">",command=lambda:upper(),height=1,width=1).place(x=38,y=19)
tkinter.Button(root,text="<",command=lambda:down(),height=1,width=1).place(x=2,y=19)

tkinter.Button(root,text="閉じる",command=lambda:close(root),width=14).place(x=250,y=350)

path="cointoss.xlsx"
for i in glob.glob("*.xlsx"):
    if i==path:
        a=True
    else:
        a=False
        
if not a ==True:
    GY=openpyxl.Workbook()
    sheet=GY.active
    sheet.title="GBP_JPY"
    GY.save(path)
else:
    GY=openpyxl.load_workbook(path)
    sheet=GY.active
    GY.save(path)
    
'''for i in range(len(text2)):
    for j in range(5):
        sheet[chr(ord("H")+i)+str(5*(int(edit.get())-1)+j+1)]=text2[i]
        GY.save(path)'''

tkinter.Button(root,text="入力",command=lambda:btn_click(),width=14).place(x=50,y=300)
for i in range(len(text2)):
    tkinter.Label(root,text=text2[i]).place(x=350,y=i*50+50)
tkinter.Button(root,text="詳細変更",command=lambda:new_win(),width=14).place(x=250,y=300)
tkinter.Button(root,text="リセット",command=lambda:resetting(bln),width=14).place(x=150,y=300)

#関数
def btn_click():
    n=int(edit.get())
    for i in range(20):
        if 0<=i<=4 or 10<=i<=14:
            if bln[i].get() ==True and bln[i+5].get() ==True:
                messagebox.showerror("Error","check your checkbox")
                sys.exit()
    if n==0:
        messagebox.showerror("Error","input correct number")
        sys.exit()

        

    for i in range(len(text2)):
        sheet[chr(ord("H")+i)+str(5*(n-1)+1)]=text2[i]               
        GY.save(path)
        
    for i in range(4):
        for j in range(5):
            if 0<=(5*i+j)<5:
                if bln[i*5+j].get()==True and bln[i*5+j+5].get()==False:
                    sheet[chr(ord("A")+i)+str(5*(n-1)+j+1)]=1
                    sheet[chr(ord("A")+i+1)+str(5*(n-1)+j+1)]=-1
                elif bln[i*5+j].get()==False and bln[i*5+j+5].get()==True:
                    sheet[chr(ord("A")+i)+str(5*(n-1)+j+1)]=-1
                    sheet[chr(ord("A")+i+1)+str(5*(n-1)+j+1)]=1
                else:
                    sheet[chr(ord("A")+i)+str(5*(n-1)+j+1)]=0
                    sheet[chr(ord("A")+i+1)+str(5*(n-1)+j+1)]=0
                    
            if 10<=(i*5+j)<15:
                if bln[i*5+j].get()==True:
                    sheet[chr(ord("A")+i)+str(5*(n-1)+j+1)]=text2[3]
                else:
                    sheet[chr(ord("A")+i)+str(5*(n-1)+j+1)]=0
            if 15<=(i*5+j):
                if bln[i*5+j].get()==True:
                    sheet[chr(ord("A")+i)+str(5*(n-1)+j+1)]=text2[5]
                else:
                    sheet[chr(ord("A")+i)+str(5*(n-1)+j+1)]=0
    
    GY.save(path)
    
    observe=sheet.iter_cols(min_row=1,min_col=1,max_col=1)
    #reserve=sheet.iter_cols(min_row=1,min_col=2,max_col=2)
    win=sheet.iter_cols(min_row=1,min_col=3,max_col=3)
    lose=sheet.iter_cols(min_row=1,min_col=4,max_col=4)
    
    cob=count_observe(observe)
    
    win_data=count_data(win)
    lose_data=count_data(lose)
    
    p_t_data=np.add(win_data,lose_data)
    print("総損益(pips)："+str(np.sum(p_t_data)))
    print("総損益(金額)："+str(np.sum(p_t_data)*text2[1]*100))
    
    redata,count_zero=clean_data(p_t_data)
    
    ds_coin=data_sum(cob)
    ds_pips=data_sum(redata)
    
    co_coin,co_pips=generate(len(ds_coin))
    co_ds_coin=data_sum(co_coin)
    co_ds_pips=data_sum(co_pips)
    
    times=[i for i in range(1,len(ds_coin)+1)]
    
    fig=plt.figure()
    ax1 = fig.add_subplot(1, 2, 1)
    ax1.plot(times, ds_coin, marker="o", color = "red", linestyle = "--")
    ax1.grid()
    #ax1.plot(times,co_ds_coin, marker="v", color = "blue", linestyle = ":")

    ax2 = fig.add_subplot(1, 2, 2)
    ax2.plot(times,ds_pips, marker="o", color = "red", linestyle = "--")
    ax2.plot(times,co_ds_pips, marker="v", color = "blue", linestyle = ":")
    ax2.grid()
    plt.show()
    
    
def new_win():
    root2=tkinter.Tk()
    root2.geometry("300x250")
    root2.title("詳細変更")

    tkinter.Label(root2,text="変更前").place(x=100,y=20)
    tkinter.Label(root2,text="変更後").place(x=200,y=20)
    
    b=50
    for i in range(3):
        a=50
        for j in range(2):
            tkinter.Label(root2,text=text2[i*2+j]).place(x=a,y=b)
            a+=50
        b+=50
    new_lot=tkinter.IntVar()
    edit1=tkinter.ttk.Entry(root2,textvariable=new_lot)
    edit1.place(x=150,y=50)
    new_profit=tkinter.IntVar()
    edit2=tkinter.ttk.Entry(root2,textvariable=new_profit)
    edit2.place(x=150,y=100)
    new_losscut=tkinter.IntVar()
    edit3=tkinter.ttk.Entry(root2,textvariable=new_losscut)
    edit3.place(x=150,y=150)

    tkinter.Button(root2,text="保存",command=lambda:save(root2,edit1.get(),edit2.get(),edit3.get()),width=14).place(x=100,y=200)

    root2.mainloop()
    
def save(root2,a,b,c):
    text2[1]=int(a)
    text2[3]=int(b)
    text2[5]=int(c)
    root2.destroy()


def count_observe(data):
    count_zero=0
    count_re=0
    count_ob=0
    cob=[]

    for i in data:
        for j in i:
            if j.value==0:
                count_zero+=1
            elif j.value==-1:
                cob.append(j.value)
                count_re+=1
            else:
                cob.append(j.value)
                count_ob+=1
                
    print("試行回数："+str(count_ob+count_re)+"回")
    print("取り逃し："+str(count_zero)+"回")
    print("表：{:.3g}".format(count_ob/(count_re+count_ob)*100)+"%")    
    return cob
    
def count_data(data):
    a=np.array([])
    for i in data:
        for j in i:
            a=np.append(a,j.value)
    return a

def clean_data(data):
    redata=np.array([])
    count_zero=0
    for i in data:
        if not i==0:
            redata=np.append(redata,i)
        else:
            count_zero+=1
    return redata,count_zero

def data_sum(data):
    a=np.array([])
    sum=0
    for i in data:
        sum=sum+i
        a=np.append(a,sum)
    return a
    
def generate(n):
    data1=np.array([])
    data2=np.array([])
    for i in range(n):
        if i%40==0:
            data1=np.append(data1,1)
            data2=np.append(data2,-50)
        elif i%2==0:
            data1=np.append(data1,1)
            data2=np.append(data2,60)
        else:
            data1=np.append(data1,-1)
            data2=np.append(data2,-50)
    return data1,data2

def resetting(bln):
    for i in range(len(bln)):
        bln[i].set(False)

def upper():
    a=int(edit.get())
    edit.delete(0,tkinter.END)
    edit.insert(tkinter.END,str(a+1))
        
def down():
    a=int(edit.get())
    edit.delete(0,tkinter.END)
    edit.insert(tkinter.END,str(a-1))
    
def close(root):
    root.destroy()

def update():
    root.after(1000,update)

root.after(1000,update)
root.mainloop()